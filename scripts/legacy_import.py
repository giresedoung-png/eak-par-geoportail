#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
legacy_import.py — Intégration des données papier historiques (pré-Kobo)
==========================================================================

Calé précisément sur la structure réelle de vos deux fichiers :
  - Base_donnees_Verification_PAP_EAK.xlsx  (feuille "Base_donnees", en-têtes ligne 4)
    -> volet "Vérification décret"
  - BDD_Enquete_PAR_EAK.xlsx  (feuille "BDD_Menages", codes ligne 2 / libellés ligne 3,
    données à partir de la ligne 4)
    -> volet "Identification nouveaux PAP"

Particularités gérées :
  - Le masque "nouveau" a ses colonnes Q0_4_Longitude / Q0_4_Latitude
    EFFECTIVEMENT INVERSÉES par rapport à leur libellé (vérifié empiriquement :
    les valeurs de "Longitude" sont en réalité des latitudes ~2-3° et
    inversement) — corrigé automatiquement par détection de plage plausible.
  - Les coordonnées peuvent être en degrés-minutes-secondes texte
    (ex: "02°48'51,62''") ou en entier brut ponctuel — les deux sont gérés.
  - Aucune colonne de référence photo n'étant renseignée dans les masques,
    l'association aux photos NoteCam se fait par correspondance floue de
    noms (nom du PAP + village) sur les fichiers renommés manuellement,
    et par proximité GPS/date pour les photos encore nommées par horodatage.

Dépendances : pip install openpyxl Pillow
"""
import csv
import json
import re
import unicodedata
import difflib
from pathlib import Path
from datetime import datetime

import openpyxl
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS

# =============================================================================
# CONFIGURATION
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent.parent
LEGACY_DIR = BASE_DIR / "legacy"
MASQUE_DECRET_XLSX = LEGACY_DIR / "Base_donnees_Verification_PAP_EAK.xlsx"
MASQUE_NOUVEAU_XLSX = LEGACY_DIR / "BDD_Enquete_PAR_EAK.xlsx"
NOTECAM_PHOTOS_DIR = LEGACY_DIR / "Images de terrain"
OUTPUT_DIR = BASE_DIR / "site" / "data"
OUTPUT_MEDIA_DIR = BASE_DIR / "site" / "media" / "legacy"

# Plage plausible de latitude/longitude pour le corridor RN17 (Sud Cameroun),
# utilisée pour détecter et corriger l'inversion lon/lat du masque "nouveau".
LAT_RANGE = (1.5, 4.5)     # degrés Nord
LON_RANGE = (8.5, 13.0)    # degrés Est

# Date de bascule vers Kobo : toute photo horodatée à partir de cette date
# doit être signalée comme risque potentiel de double collecte (papier + Kobo
# en parallèle), plutôt que d'être traitée comme une donnée papier "normale".
KOBO_LAUNCH_DATE = "20260721"  # AAAAMMJJ — à ajuster si besoin

# Seuil de similarité (0-1) pour la correspondance floue nom de PAP <-> nom de photo
NAME_MATCH_THRESHOLD = 0.55


# =============================================================================
# NORMALISATION DE TEXTE (pour comparaison floue de noms)
# =============================================================================

def normalize_text(s):
    if not s:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", " ", s).strip().lower()
    return s


def similarity(a, b):
    """Similarité combinant recouvrement de tokens (Jaccard), similarité de
    caractères (ratio de séquence) et taux de couverture (les mots du nom
    'a' se retrouvent-ils tous dans 'b' ?) — ce dernier critère est
    essentiel pour les noms de fichiers descriptifs et longs du type
    'Habitation de MBANG ALO'O Réné (Bagyeli par alliance)', où le nom du
    PAP est noyé dans du texte additionnel qui ferait chuter le seul score
    de Jaccard/caractères."""
    na, nb = normalize_text(a), normalize_text(b)
    ta, tb = set(na.split()), set(nb.split())
    jaccard = len(ta & tb) / len(ta | tb) if ta and tb else 0.0
    char_ratio = difflib.SequenceMatcher(None, na, nb).ratio()
    couverture = len(ta & tb) / len(ta) if ta else 0.0
    return max(jaccard, char_ratio, couverture)


# =============================================================================
# PARSING DES COORDONNÉES (DMS texte, entier ponctuel, avec correction d'inversion)
# =============================================================================

def parse_dms_string(s):
    """Parse '02°48'51,62''' -> 2.813..."""
    m = re.match(r"(\d+)[°\s]+(\d+)['\s]+([\d,\.]+)", str(s).strip())
    if not m:
        return None
    deg, minute, sec = m.groups()
    sec = sec.replace(",", ".")
    try:
        return float(deg) + float(minute) / 60.0 + float(sec) / 3600.0
    except ValueError:
        return None


def parse_coordinate_value(raw):
    """Convertit une valeur brute de coordonnée (texte DMS ou nombre) en degrés
    décimaux, quel que soit le format rencontré dans le masque."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        # Valeur numérique déjà en degrés décimaux si dans une plage plausible
        if LAT_RANGE[0] <= raw <= LON_RANGE[1]:
            return float(raw)
        return None  # format non identifié (ex: encodage propriétaire) -> ignoré, pas de valeur inventée
    return parse_dms_string(raw)


def resolve_lat_lon(raw_col_longitude, raw_col_latitude):
    """Retourne (lat, lon) corrects, en détectant si les colonnes du masque
    sont inversées par rapport à leur libellé (vérifié empiriquement sur ce
    projet : c'est le cas)."""
    val_a = parse_coordinate_value(raw_col_longitude)  # colonne nommée "Longitude"
    val_b = parse_coordinate_value(raw_col_latitude)   # colonne nommée "Latitude"
    if val_a is None or val_b is None:
        return None, None

    a_is_lat = LAT_RANGE[0] <= val_a <= LAT_RANGE[1]
    b_is_lon = LON_RANGE[0] <= val_b <= LON_RANGE[1]
    a_is_lon = LON_RANGE[0] <= val_a <= LON_RANGE[1]
    b_is_lat = LAT_RANGE[0] <= val_b <= LAT_RANGE[1]

    if a_is_lat and b_is_lon:
        # La colonne "Longitude" contient en fait la latitude, et vice versa (cas détecté)
        return val_a, val_b
    if a_is_lon and b_is_lat:
        # Les colonnes correspondent bien à leur libellé
        return val_b, val_a
    return None, None  # ambigu -> on ne devine pas, on laisse vide pour verification manuelle


# =============================================================================
# EXIF GPS (photos NoteCam horodatées, sans nom explicite)
# =============================================================================

def _to_degrees(value):
    def _component(v):
        if isinstance(v, tuple) and len(v) == 2:
            return v[0] / v[1]
        return float(v)
    d, m, s = value
    return _component(d) + _component(m) / 60.0 + _component(s) / 3600.0


def read_exif_gps(photo_path):
    try:
        img = Image.open(photo_path)
        exif_raw = img._getexif()
        if not exif_raw:
            return None, None
        exif = {TAGS.get(k, k): v for k, v in exif_raw.items()}
        gps_info = exif.get("GPSInfo")
        if not gps_info:
            return None, None
        gps = {GPSTAGS.get(k, k): v for k, v in gps_info.items()}
        lat = _to_degrees(gps["GPSLatitude"])
        if gps.get("GPSLatitudeRef") == "S":
            lat = -lat
        lon = _to_degrees(gps["GPSLongitude"])
        if gps.get("GPSLongitudeRef") == "W":
            lon = -lon
        if lat != lat or lon != lon:  # test NaN (certains appareils écrivent des NaN si le GPS a échoué)
            return None, None
        return lat, lon
    except Exception:
        return None, None


def build_photo_index():
    """Indexe les photos du dossier NoteCam : GPS EXIF si disponible, et nom
    de fichier normalisé pour la correspondance floue avec les PAP nommés."""
    index = []
    if not NOTECAM_PHOTOS_DIR.exists():
        print(f"  ⚠ Dossier photos introuvable : {NOTECAM_PHOTOS_DIR}")
        return index
    for photo_path in NOTECAM_PHOTOS_DIR.rglob("*"):
        if photo_path.suffix.lower() not in (".jpg", ".jpeg", ".png"):
            continue
        lat, lon = read_exif_gps(photo_path)
        # Les photos renommées manuellement (pas un simple horodatage
        # AAAAMMJJ_HHMMSS) sont candidates à la correspondance par nom.
        is_timestamp_name = bool(re.match(r"^\d{8}_\d{6}$", photo_path.stem))
        index.append({
            "path": photo_path, "lat": lat, "lon": lon,
            "nom_normalise": normalize_text(photo_path.stem),
            "is_timestamp_name": is_timestamp_name,
        })
    return index


# Mots-clés (déjà normalisés : sans accent, minuscule) pour catégoriser
# automatiquement chaque photo selon le vocabulaire de terrain observé dans
# vos fichiers réels (habitation, cultures, tombes, CNI, infrastructures...).
CATEGORIES_PHOTO = [
    ("piece_identite", ["cni", "carte identite", "piece identite"]),
    ("tombe", ["tombe", "cimetiere", "sepulture"]),
    ("infrastructure_eau", ["puit", "forage", "borne fontaine", "impluvium"]),
    ("cultures", ["culture", "parcelle cultiv", "cacao", "cour cultiv", "arbre fruitier",
                  "plantation", "champ"]),
    ("infrastructure_communautaire", ["stelle", "eglise", "chefferie"]),
    ("batiment_concession", ["habitation", "concession", "domaine", "batiment"]),
]


def categorize_photo(filename_stem):
    """Déduit la catégorie d'une photo à partir de mots-clés dans son nom de
    fichier. Par défaut (aucun mot-clé trouvé) : portrait/photo générale du
    répondant. La comparaison ignore les pluriels simples (arbre/arbres)."""
    s = normalize_text(filename_stem)
    s_sing = re.sub(r"(\w)s\b", r"\1", s)  # "arbres fruitiers" -> "arbre fruitier"
    for categorie, mots_cles in CATEGORIES_PHOTO:
        for mot in mots_cles:
            mot_sing = re.sub(r"(\w)s\b", r"\1", mot)
            if mot in s or mot_sing in s_sing:
                return categorie
    return "photo_repondant"


def match_photos_to_paps(all_rows, photo_index):
    """Pour chaque photo nommée (non horodatée), cherche la PAP dont le nom
    correspond le mieux — une même PAP peut ainsi récupérer PLUSIEURS photos
    (habitation, cultures, tombes, CNI...), contrairement à une simple
    correspondance 1 photo <-> 1 PAP."""
    assignments = {}  # index de ligne (dans all_rows) -> liste de (photo_entry, score)
    unmatched = []

    for entry in photo_index:
        if entry["is_timestamp_name"]:
            unmatched.append(entry)
            continue
        best_idx, best_score = None, 0.0
        for i, row in enumerate(all_rows):
            nom = row.get("nom_prenom_cm") or ""
            village = row.get("village") or ""
            score = similarity(nom, entry["path"].stem)
            score = max(score, similarity(f"{village} {nom}", entry["path"].stem))
            if score > best_score:
                best_score, best_idx = score, i
        if best_idx is not None and best_score >= NAME_MATCH_THRESHOLD:
            assignments.setdefault(best_idx, []).append((entry, best_score))
        else:
            unmatched.append(entry)

    return assignments, unmatched


def find_matching_photo(nom_pap, village, photo_index, used_photos):
    """Cherche la meilleure correspondance floue nom+village parmi les photos
    NON horodatées (donc explicitement renommées avec un nom de PAP)."""
    best, best_score = None, 0.0
    candidate_text = f"{village or ''} {nom_pap or ''}"
    for entry in photo_index:
        if entry["is_timestamp_name"] or entry["path"] in used_photos:
            continue
        score = similarity(candidate_text, entry["path"].stem)
        # Bonus si le nom du PAP seul correspond bien, même sans le village
        score = max(score, similarity(nom_pap, entry["path"].stem))
        if score > best_score:
            best, best_score = entry, score
    if best and best_score >= NAME_MATCH_THRESHOLD:
        return best
    return None


# =============================================================================
# LECTURE DES MASQUES (structure réelle)
# =============================================================================

def read_masque_decret():
    if not MASQUE_DECRET_XLSX.exists():
        print(f"  (masque décret absent, ignoré : {MASQUE_DECRET_XLSX.name})")
        return []
    wb = openpyxl.load_workbook(MASQUE_DECRET_XLSX, data_only=True)
    ws = wb["Base_donnees"]
    headers = [c.value for c in ws[4]]
    rows = []
    for excel_row in ws.iter_rows(min_row=5, values_only=True):
        raw = dict(zip(headers, excel_row))
        # Une ligne est valide si au moins le nom du PAP est renseigné
        if not raw.get("Nom_prenom_PAP"):
            continue
        rows.append({
            "type_fiche": "decret",
            "departement": raw.get("Departement"),
            "arrondissement": raw.get("Arrondissement"),
            "village": raw.get("Village"),
            "date_enquete": raw.get("Date_enquete"),
            "nom_prenom_cm": raw.get("Nom_prenom_PAP"),
            "pk": None,           # non collecté sur ce volet papier
            "latitude": None,     # non collecté sur ce volet papier
            "longitude": None,
            "decret_ref": raw.get("Decret_ref_N"),
            "nature_biens_affectes": raw.get("Nature_biens_affectes"),
            "presence_decret_original": raw.get("Presence_decret_orig"),
            "requete_reclamation": raw.get("Requete_reclamation"),
        })
    print(f"  {len(rows)} PAP lue(s) dans le masque décret")
    return rows


def read_masque_nouveau():
    if not MASQUE_NOUVEAU_XLSX.exists():
        print(f"  (masque nouveau absent, ignoré : {MASQUE_NOUVEAU_XLSX.name})")
        return []
    wb = openpyxl.load_workbook(MASQUE_NOUVEAU_XLSX, data_only=True)
    ws = wb["BDD_Menages"]
    codes = [c.value for c in ws[2]]

    def col(code):
        return codes.index(code) + 1 if code in codes else None

    c_code_fiche = col("Q0_3_CodeFiche")
    c_lon = col("Q0_4_Longitude")
    c_lat = col("Q0_4_Latitude")
    c_commune = col("Q0_5_Commune")
    c_village = col("Q0_5_Village")
    c_pk = col("Q0_5_PK")
    c_enqueteur = col("Q0_6_Enqueteur")
    c_nom_cm = col("QII_1_NomPrenomsCM")

    rows = []
    for r in range(4, ws.max_row + 1):
        nom_cm = ws.cell(row=r, column=c_nom_cm).value if c_nom_cm else None
        if not nom_cm:
            continue
        raw_lon = ws.cell(row=r, column=c_lon).value if c_lon else None
        raw_lat = ws.cell(row=r, column=c_lat).value if c_lat else None
        lat, lon = resolve_lat_lon(raw_lon, raw_lat)
        rows.append({
            "type_fiche": "nouveau",
            "departement": None,
            "arrondissement": None,
            "village": ws.cell(row=r, column=c_village).value if c_village else None,
            "commune": ws.cell(row=r, column=c_commune).value if c_commune else None,
            "date_enquete": None,
            "nom_prenom_cm": nom_cm,
            "pk": ws.cell(row=r, column=c_pk).value if c_pk else None,
            "latitude": lat,
            "longitude": lon,
            "code_fiche": ws.cell(row=r, column=c_code_fiche).value if c_code_fiche else None,
            "nom_enqueteur": ws.cell(row=r, column=c_enqueteur).value if c_enqueteur else None,
        })
    print(f"  {len(rows)} PAP lue(s) dans le masque nouveau (sur {ws.max_row} lignes explorées)")
    return rows


# =============================================================================
# TRAITEMENT PRINCIPAL
# =============================================================================

def process_legacy_data():
    print("Lecture des masques de saisie...")
    rows_decret = read_masque_decret()
    rows_nouveau = read_masque_nouveau()
    all_rows = rows_decret + rows_nouveau

    print("Indexation des photos NoteCam...")
    photo_index = build_photo_index()
    n_named = sum(1 for p in photo_index if not p["is_timestamp_name"])
    n_geo = sum(1 for p in photo_index if p["lat"] is not None)
    print(f"  {len(photo_index)} photo(s) au total | {n_named} renommée(s) (nom de PAP) "
          f"| {n_geo} avec GPS EXIF exploitable")

    assignments, unmatched_photos = match_photos_to_paps(all_rows, photo_index)

    OUTPUT_MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    main_rows = []
    manifest_rows = []
    n_photos_liees = 0

    for i, row in enumerate(all_rows):
        uuid = f"legacy-{i:05d}"
        lat, lon = row.get("latitude"), row.get("longitude")
        photos_pour_cette_pap = sorted(assignments.get(i, []), key=lambda x: -x[1])

        # Si la PAP n'a pas de coordonnées propres, utiliser le GPS de la
        # première photo associée qui en dispose (à défaut de mieux).
        if lat is None:
            for entry, _score in photos_pour_cette_pap:
                if entry["lat"] is not None:
                    lat, lon = entry["lat"], entry["lon"]
                    break

        main_rows.append({
            "_uuid": uuid, "_id": uuid,
            "_submission_time": row.get("date_enquete"),
            "type_fiche": row.get("type_fiche"),
            "source": "papier_historique",
            "troncon": None,
            "departement": row.get("departement"),
            "arrondissement": row.get("arrondissement"),
            "village": row.get("village"),
            "date_enquete": row.get("date_enquete"),
            "nom_prenom_cm": row.get("nom_prenom_cm"),
            "pk": row.get("pk"),
            "latitude": lat, "longitude": lon,
            "altitude": None, "precision_gps": None,
            "nb_photos_associees": len(photos_pour_cette_pap),
        })

        # Compteur par catégorie pour nommer les fichiers de façon lisible
        # s'il y a plusieurs photos de la même catégorie pour une même PAP
        compteur_categorie = {}
        for entry, score in photos_pour_cette_pap:
            categorie = categorize_photo(entry["path"].stem)
            compteur_categorie[categorie] = compteur_categorie.get(categorie, 0) + 1
            suffixe = f"_{compteur_categorie[categorie]}" if compteur_categorie[categorie] > 1 else ""
            dest = OUTPUT_MEDIA_DIR / f"{uuid}__{categorie}{suffixe}{entry['path'].suffix.lower()}"
            dest.write_bytes(entry["path"].read_bytes())
            manifest_rows.append({
                "_uuid": uuid, "_id": uuid, "champ": categorie,
                "chemin_local": str(dest.relative_to(BASE_DIR / "site")),
                "photo_source": entry["path"].name,
                "score_correspondance": round(score, 2),
            })
            n_photos_liees += 1

    n_paps_avec_photo = sum(1 for r in main_rows if r["nb_photos_associees"] > 0)
    print(f"\n{n_photos_liees} photo(s) liée(s) à {n_paps_avec_photo} PAP "
          f"(sur {len(all_rows)} PAP au total, {n_named} photos nommées disponibles).")

    unmatched = unmatched_photos
    unmatched_post_kobo = [p for p in unmatched if p["path"].stem[:8].isdigit() and p["path"].stem[:8] >= KOBO_LAUNCH_DATE]
    unmatched_pre_kobo = [p for p in unmatched if p not in unmatched_post_kobo]

    if unmatched_post_kobo:
        print(f"\n🔴 ATTENTION — {len(unmatched_post_kobo)} photo(s) horodatée(s) à partir du "
              f"{KOBO_LAUNCH_DATE} (date de bascule vers Kobo) et NON associée(s) à une PAP "
              f"du masque papier :")
        for p in unmatched_post_kobo:
            geo = f"(GPS: {p['lat']:.5f}, {p['lon']:.5f})" if p["lat"] else "(sans GPS)"
            print(f"    - {p['path'].name}  {geo}")
        print("   ⚠ Ces photos coïncident avec le début de la collecte Kobo : vérifiez qu'elles "
              "ne correspondent pas à une PAP DÉJÀ saisie via le formulaire Kobo (même personne "
              "photographiée deux fois), pour éviter un double comptage sur le géoportail.")

    if unmatched_pre_kobo:
        print(f"\n⚠ {len(unmatched_pre_kobo)} autre(s) photo(s) NON associée(s) automatiquement "
              f"(sans nom exploitable ou similarité insuffisante avec un nom de PAP) :")
        for p in unmatched_pre_kobo[:20]:
            reason = "horodatage seul, pas de nom exploitable" if p["is_timestamp_name"] else "nom sans correspondance suffisante dans les masques"
            geo = f"(GPS: {p['lat']:.5f}, {p['lon']:.5f})" if p["lat"] else "(sans GPS)"
            print(f"    - {p['path'].name}  {geo}  [{reason}]")

    return main_rows, manifest_rows


def export_csv(rows, path):
    if not rows:
        return
    keys = []
    for r in rows:
        for k in r:
            if k not in keys:
                keys.append(k)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)


def main():
    LEGACY_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    main_rows, manifest_rows = process_legacy_data()

    export_csv(main_rows, OUTPUT_DIR / "legacy_principal.csv")
    export_csv(manifest_rows, OUTPUT_DIR / "legacy_manifest_media.csv")

    n_geo = sum(1 for r in main_rows if r.get("latitude") is not None)
    print(f"\n{len(main_rows)} PAP historique(s) au total, dont {n_geo} géolocalisée(s), "
          f"{len(manifest_rows)} photo(s) intégrée(s).")
    print("Prochaine étape : lancer merge_sources.py.")


if __name__ == "__main__":
    main()
