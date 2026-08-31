#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
analyse_emprise.py — Analyse des PAP situées dans l'emprise réelle de la RN17
================================================================================

Réutilise le PK et la bande d'emprise déjà calculés par calcul_pk_lineaire.py
(site/data/pk_calcule.csv — colonne 'bande_emprise', renseignée uniquement
pour les PAP couvertes par le levé topographique réel, seule méthode assez
précise pour ce classement) et produit :

  site/data/PAP_Zone_Tampon_Par_Village.xlsx — les PAP dont le bâti se
  trouve dans la zone tampon (6,5 à 11,5 m de l'axe), organisées par feuille
  Excel séparée par village.

À exécuter après calcul_pk_lineaire.py (qui produit pk_calcule.csv).
"""
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "site" / "data"

PK_CALCULE_CSV = DATA_DIR / "pk_calcule.csv"
OUT_XLSX_ZONE_TAMPON = DATA_DIR / "PAP_Zone_Tampon_Par_Village.xlsx"

LABELS_BANDE = {
    "chaussee": "Chaussée (0-3,5 m)",
    "trottoir": "Trottoir (3,5-5 m)",
    "rigole": "Rigole (5-6,5 m)",
    "zone_tampon": "Zone tampon (6,5-11,5 m)",
    "hors_emprise": "Hors emprise (> 11,5 m)",
}


def main():
    if not PK_CALCULE_CSV.exists():
        print(f"⚠ {PK_CALCULE_CSV.name} introuvable — lancez d'abord calcul_pk_lineaire.py.")
        return

    with open(PK_CALCULE_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    couvertes = [r for r in rows if r.get("bande_emprise")]
    print(f"{len(rows)} PAP au total dans pk_calcule.csv, dont {len(couvertes)} couvertes "
          f"par le levé topographique réel (seule méthode fiable pour le classement par bande).")

    compte_bandes = {}
    for r in couvertes:
        compte_bandes[r["bande_emprise"]] = compte_bandes.get(r["bande_emprise"], 0) + 1
    print("\nRépartition par bande d'emprise :")
    for bande, label in LABELS_BANDE.items():
        print(f"    {label} : {compte_bandes.get(bande, 0)}")

    zone_tampon = [r for r in couvertes if r["bande_emprise"] == "zone_tampon"]
    if not zone_tampon:
        print("\nAucune PAP en zone tampon détectée — pas de fichier Excel généré.")
        return

    def village_de(r):
        return r.get("village") or "Non renseigné"

    villages = sorted(set(village_de(r) for r in zone_tampon))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    entete = ["Nom et prénom", "Type de fiche", "Origine", "PK (km)",
              "Distance à l'axe (m)", "Latitude", "Longitude"]

    for village in villages:
        nom_feuille = village[:31]  # limite Excel pour les noms de feuille
        ws = wb.create_sheet(nom_feuille)
        ws.append([f"PAP EN ZONE TAMPON — {village}"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(entete))
        ws["A1"].font = Font(bold=True, size=13)
        ws.append(entete)
        for cell in ws[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C0392B")

        lignes_village = sorted(
            [r for r in zone_tampon if village_de(r) == village],
            key=lambda r: float(r["pk_calcule_km"])
        )
        for r in lignes_village:
            nom = r.get("nom_prenom_cm") or "Nom non renseigné"
            type_label = "Vérification décret" if r.get("type_fiche") == "decret" else "Nouveau PAP potentiel"
            origine = "Papier historique" if r.get("source") == "papier_historique" else "Kobo"
            ws.append([
                nom, type_label, origine, float(r["pk_calcule_km"]),
                float(r["distance_axe_m"]), r.get("latitude"), r.get("longitude"),
            ])
        for col_idx in range(1, len(entete) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 22

    # Feuille de synthèse en première position
    ws_synthese = wb.create_sheet("SYNTHÈSE", 0)
    ws_synthese.append(["SYNTHÈSE — PAP EN ZONE TAMPON DE L'EMPRISE RN17 (6,5 à 11,5 m de l'axe)"])
    ws_synthese.merge_cells("A1:B1")
    ws_synthese["A1"].font = Font(bold=True, size=13)
    ws_synthese.append([])
    ws_synthese.append(["Village", "Nombre de PAP en zone tampon"])
    for cell in ws_synthese[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0392B")
    for village in villages:
        nb = sum(1 for r in zone_tampon if village_de(r) == village)
        ws_synthese.append([village, nb])
    ws_synthese.append(["TOTAL", len(zone_tampon)])
    ws_synthese["A" + str(ws_synthese.max_row)].font = Font(bold=True)
    ws_synthese["B" + str(ws_synthese.max_row)].font = Font(bold=True)
    ws_synthese.column_dimensions["A"].width = 25
    ws_synthese.column_dimensions["B"].width = 30

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX_ZONE_TAMPON)
    print(f"\n  ✓ {OUT_XLSX_ZONE_TAMPON.name} : {len(zone_tampon)} PAP en zone tampon, "
          f"réparties sur {len(villages)} village(s) (+ 1 feuille de synthèse)")


if __name__ == "__main__":
    main()
